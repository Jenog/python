import sys
import pyodbc
import re
import decimal
import openpyxl.styles.alignment
import time
import ctypes

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt,QCoreApplication
from wcwidth import wcswidth
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font




# 쿼리 문

select_all_q_Y =  "select * from fillib.jpypf where jpdtcy='2019' and jpfact='Y'"
select_all_q_T =  "select * from fillib.jpypf where jpdtcy='2019' and jpfact='T'"
select_all_q_W =  "select * from fillib.jpypf where jpdtcy='2019' and jpfact='W'"

number_of_people_q_Y = "select count(*) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y'  " # 전체 인원수 양산
number_of_people_q_T = "select count(*) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T'  " # 전체 인원수 대전
number_of_people_q_W = "select count(*) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W'  " # 전체 인원수 음성

total_of_gup_q_Y_6 = "select sum(jppd01+jppd02+jppd03+jppd04+jppd05+jppd06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y'  " # 급여총액 (1-6월) 양산
total_of_gup_q_T_6 = "select sum(jppd01+jppd02+jppd03+jppd04+jppd05+jppd06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T'  " # 급여총액 (1-6월) 대전
total_of_gup_q_W_6 = "select sum(jppd01+jppd02+jppd03+jppd04+jppd05+jppd06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W'  " # 급여총액 (1-6월) 음성

total_of_sang_q_Y_6 = "select sum(jpbs01+jpbs02+jpbs03+jpbs04+jpbs05+jpbs06+jpbd01+jpbd02+jpbd03+jpbd04+jpbd05+jpbd06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y'  " # 상여총액 (1-6월) 양산
total_of_sang_q_T_6 = "select sum(jpbs01+jpbs02+jpbs03+jpbs04+jpbs05+jpbs06+jpbd01+jpbd02+jpbd03+jpbd04+jpbd05+jpbd06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T'  " # 상여총액 (1-6월) 대전
total_of_sang_q_W_6 = "select sum(jpbs01+jpbs02+jpbs03+jpbs04+jpbs05+jpbs06+jpbd01+jpbd02+jpbd03+jpbd04+jpbd05+jpbd06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W'  " # 상여총액 (1-6월) 음성

total_of_bi_q_Y_6 = "select sum(jpbi01+jpbi02+jpbi03+jpbi04+jpbi05+jpbi06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y' " #비과세 총액 (1-6월) 양산
total_of_bi_q_T_6 = "select sum(jpbi01+jpbi02+jpbi03+jpbi04+jpbi05+jpbi06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T' " #비과세 총액 (1-6월) 대전
total_of_bi_q_W_6 = "select sum(jpbi01+jpbi02+jpbi03+jpbi04+jpbi05+jpbi06) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W' " #비과세 총액 (1-6월) 음성


total_of_gup_q_Y_12 = "select sum(jppd07+jppd08+jppd09+jppd10+jppd11+jppd12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y'  " # 급여총액 (7-12월) 양산
total_of_gup_q_T_12 = "select sum(jppd07+jppd08+jppd09+jppd10+jppd11+jppd12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T'  " # 급여총액 (7-12월) 대전
total_of_gup_q_W_12 = "select sum(jppd07+jppd08+jppd09+jppd10+jppd11+jppd12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W'  " # 급여총액 (7-12월) 음성

total_of_sang_q_Y_12 = "select sum(jpbs07+jpbs08+jpbs09+jpbs10+jpbs11+jpbs12+jpbd07+jpbd08+jpbd09+jpbd10+jpbd11+jpbd12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y'  " # 상여총액 (7-12월) 양산
total_of_sang_q_T_12 = "select sum(jpbs07+jpbs08+jpbs09+jpbs10+jpbs11+jpbs12+jpbd07+jpbd08+jpbd09+jpbd10+jpbd11+jpbd12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T'  " # 상여총액 (7-12월) 대전
total_of_sang_q_W_12 = "select sum(jpbs07+jpbs08+jpbs09+jpbs10+jpbs11+jpbs12+jpbd07+jpbd08+jpbd09+jpbd10+jpbd11+jpbd12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W'  " # 상여총액 (7-12월) 음성

total_of_bi_q_Y_12 = "select sum(jpbi07+jpbi08+jpbi09+jpbi10+jpbi11+jpbi12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'Y' " #비과세 총액 (7-12월) 양산
total_of_bi_q_T_12 = "select sum(jpbi07+jpbi08+jpbi09+jpbi10+jpbi11+jpbi12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'T' " #비과세 총액 (7-12월) 대전
total_of_bi_q_W_12 = "select sum(jpbi07+jpbi08+jpbi09+jpbi10+jpbi11+jpbi12) from fillib.jpypf where jpdtcy = '2019' and jpfact = 'W' " #비과세 총액 (7-12월) 음성


join_for_phone_Y = "select j.jphnme, j.jpfact, j.jpjnyy, j.jpjseq, i.ishptl from fillib.jpypf j, fillib.isapf i where j.jpfact = i.isfact and j.jpjnyy=i.isjnyy and j.jpjseq=i.isjseq and j.jpdtcy='2019' and j.jpfact='Y'  " #휴대폰번호 양산
join_for_phone_T = "select j.jphnme, j.jpfact, j.jpjnyy, j.jpjseq, i.ishptl from fillib.jpypf j, fillib.isapf i where j.jpfact = i.isfact and j.jpjnyy=i.isjnyy and j.jpjseq=i.isjseq and j.jpdtcy='2019' and j.jpfact='T'  " #휴대폰번호 대전
join_for_phone_W = "select j.jphnme, j.jpfact, j.jpjnyy, j.jpjseq, i.ishptl from fillib.jpypf j, fillib.isapf i where j.jpfact = i.isfact and j.jpjnyy=i.isjnyy and j.jpjseq=i.isjseq and j.jpdtcy='2019' and j.jpfact='W'  " #휴대폰번호 음성

select_sabon_Y = "select * from fillib.isapf where isedte=' ' and isfact = 'Y'"         # 총 재직자 사번 양산
select_sabon_T = "select * from fillib.isapf where isedte=' ' and isfact = 'T'"         # 총 재직자 사번 대전
select_sabon_W = "select * from fillib.isapf where isedte=' ' and isfact = 'W'"         # 총 재직자 사번 음성

# 공장별 변경사항

code_taxoffice_Y = '621'
code_taxoffice_W = '303'
code_taxoffice_T = '314'


id_hometax_Y = 'hanil01'
id_hometax_W = 'hanil124'
id_hometax_T = 'hanil042'

number_register_Y ='6218105572'
number_register_W ='1248531562'
number_register_T ='3068512298'

depart_name_Y ='김귀철'
depart_name_W ='김종태'
depart_name_T ='문미경'

depart_phone_Y = '055-370-6600'
depart_phone_W = '043-883-2805'
depart_phone_T = '042-934-5500'


# 파일명
filename_Y = 'C:\\pay\\SC6218105.572'
filename_T = 'C:\\pay\\SC3068512.298'
filename_W = 'C:\\pay\\SC1248531.562'




class LogInDialog1(QDialog):


    def __init__(self):
        super().__init__()
        self.initUI()


    def initUI(self):
        self.setGeometry(1100, 200, 300, 100)
        self.setWindowTitle("총무")


        label1 = QLabel("공     장  : ")
        label2 = QLabel("기     간  : ")
        label1.setAlignment(Qt.AlignCenter)
        label2.setAlignment(Qt.AlignCenter)

        font1 = label1.font()
        font1.setPointSize(12)

        font1 = label2.font()
        font1.setFamily('Times New Roman')


        font2 = label2.font()
        font2.setPointSize(12)

        font2 = label2.font()
        font2.setFamily('Times New Roman')


        label1.setFont(font1)
        label2.setFont(font2)

        self.cbo1 = QComboBox()
        self.cbo2 = QComboBox()

        # 콤보박스에 아이템 리스트 넣기
        self.cbo1.addItems(['양산공장', '대전공장', '음성공장'])
        self.cbo2.addItems(['전반기', '후반기'])

        self.pushButton1= QPushButton("신고파일추출")
        self.pushButton1.clicked.connect(self.pushButtonClicked1)
        self.pushButton2= QPushButton("엑셀파일 추출")
        self.pushButton2.clicked.connect(self.pushButtonClicked2)


        layout = QGridLayout()
        layout.addWidget(label1, 0, 0)
        layout.addWidget(label2, 1, 0)
        layout.addWidget(self.cbo1, 0, 1)
        layout.addWidget(self.cbo2, 1, 1)
        layout.addWidget(self.pushButton1, 3, 0)
        layout.addWidget(self.pushButton2, 3, 1)

        self.setLayout(layout)

    def pushButtonClicked1(self):


        print('a')
        print(select_all_q_Y)
    # DB 연결
        connection = pyodbc.connect(driver='{iSeries Access ODBC Driver}', system="192.168.1.3", uid='don2000',pwd='h2011')
        c1 = connection.cursor()

        print(c1)

    #전체 인원수
        if self.cbo1.currentText() == '양산공장':
            c1.execute(number_of_people_q_Y)
        elif self.cbo1.currentText() == '대전공장':
            c1.execute(number_of_people_q_T)
        else :
            c1.execute(number_of_people_q_W)

        tmp = c1.fetchone()
        number_of_people_v = str(tmp[0])
        tmp =""




    #급여총액
        if self.cbo2.currentText()=='전반기':
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_gup_q_Y_6)
            elif self.cbo1.currentText() == '대전공장':
                c1.execute(total_of_gup_q_T_6)
            else :
                c1.execute(total_of_gup_q_W_6)
        else:
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_gup_q_Y_12)
            elif self.cbo1.currentText() == '대전공장':
                c1.execute(total_of_gup_q_T_12)
            else :
                c1.execute(total_of_gup_q_W_12)


        tmp = " ".join(re.findall("\d+",str(c1.fetchone())))  #숫자만 따로 빼서 출력
        total_of_gup_v=tmp
        tmp=""

        print(total_of_gup_v)



    #상여총액\
        if self.cbo2.currentText() =='전반기':

            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_sang_q_Y_6)
            elif self.cbo1.currentText() == '대전공장':
                c1.execute(total_of_sang_q_T_6)
            else :
                c1.execute(total_of_sang_q_W_6)
        else:
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_sang_q_Y_12)
            elif self.cbo1.currentText() == '대전공장':
                c1.execute(total_of_sang_q_T_12)
            else :
                c1.execute(total_of_sang_q_W_12)

        tmp = " ".join(re.findall("\d+",str(c1.fetchone())))
        total_of_sang_v=tmp
        tmp=""



    #급여상여 총액
        total_of_sal=str((int(total_of_gup_v)+int(total_of_sang_v)))    # total_of_gup_v     total_of_sang_v


    #비과세 총액
        if self.cbo2.currentText() =='전반기':

            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_bi_q_Y_6)
            elif self.cbo1.currentText() == '대전공장':
                c1.execute(total_of_bi_q_T_6)
            else :
                c1.execute(total_of_bi_q_W_6)
        else:
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_bi_q_Y_12)
            elif self.cbo1.currentText() == '대전공장':
                c1.execute(total_of_bi_q_T_12)
            else :
                c1.execute(total_of_bi_q_W_12)


        tmp = " ".join(re.findall("\d+",str(c1.fetchone())))
        total_of_bi_v = tmp     #total_of_bi_v
        tmp=""


        # 문자 정렬 함수
        def char_format_left(s, width, fill=' '):   # 문자 왼쪽 정렬 빈칸으로채움
            s=s.strip()                             # 1차 공백 제거
            s_width = wcswidth(s)                   # 기존 컬럼 자리수 계산
            f_width = width - s_width               # 빈칸 계산
            if f_width < 0: f_width = 0             # 자리가 딱 맞음
            fill = (fill*f_width)[:f_width]
            return s + fill


        #숫자 정렬 함수
        def num_format_right(s, width, fill='0'):   # 숫자 오른쪽 정렬 0으로 채움
            s=s.strip()                             # 1차 공백 제거
            s_width = wcswidth(s)                   # 기존 컬럼 자리수 계산
            f_width = width - s_width               # 빈칸 계산
            if f_width < 0: f_width = 0             # 자리가 딱 맞음
            fill = (fill*f_width)[:f_width]
            return fill + s


        # A,B레코드

        A1 = char_format_left('A',1)
        A2 = num_format_right('77',2)
        if self.cbo1.currentText() == '양산공장':
            A3 = char_format_left(code_taxoffice_Y,3)  #수정
        elif self.cbo1.currentText() == '대전공장':
            A3 = char_format_left(code_taxoffice_T,3)  #수정
        else:
            A3 = char_format_left(code_taxoffice_W,3)  #수정

        A4 = num_format_right('20190710',8)
        A5 = num_format_right('2',1)
        A6 = char_format_left(' ',6)

        if self.cbo1.currentText() == '양산공장':
            A7 = char_format_left('hanil01',20) #수정
        elif self.cbo1.currentText() == '대전공장':
            A7 = char_format_left('hanil042',20) #수정
        else:
            A7 = char_format_left('hanil124', 20)  #수정
        A8 = char_format_left('9000',4)

        if self.cbo1.currentText() == '양산공장':
            A9 = char_format_left(number_register_Y, 10)  # 수정
        elif self.cbo1.currentText() == '대전공장':
            A9 = char_format_left(number_register_T, 10) #수정
        else:
            A9 = char_format_left(number_register_W, 10)  #수정

        A10 = char_format_left('한일제관(주)',30)
        A11 = char_format_left('총무팀',30)

        if self.cbo1.currentText() == '양산공장':
            A12 = char_format_left(depart_name_Y, 30)  # 수정
        elif self.cbo1.currentText() == '대전공장':
            A12 = char_format_left(depart_name_T, 30)  # 수정
        else:
            A12 = char_format_left(depart_name_W, 30)  # 수정


        if self.cbo1.currentText() == '양산공장':
            A13 = char_format_left(depart_phone_Y, 15)  # 수정
        elif self.cbo1.currentText() == '대전공장':
            A13 = char_format_left(depart_phone_T, 15)  # 수정
        else:
            A13 = char_format_left(depart_phone_W, 15)  # 수정

        A14 = num_format_right('1',5)
        A15 = char_format_left(' ',25)

        record_A = A1+A2+A3+A4+A5+A6+A7+A8+A9+A10+A11+A12+A13+A14+A15

        #출력 테스트
        print(record_A)
        print(wcswidth(record_A))

        # B 레코드

        B1 = char_format_left('B',1)
        B2 = num_format_right('77',2)

        if self.cbo1.currentText() == '양산공장':
            B3 = char_format_left(code_taxoffice_Y, 3)  # 수정
        elif self.cbo1.currentText() == '대전공장':
            B3 = char_format_left(code_taxoffice_T,3) #수정
        else:
            B3 = char_format_left(code_taxoffice_W,3) #수정


        B4 = num_format_right('1',6)
        B5 = char_format_left('한일제관(주)',40)
        B6 = char_format_left('정동택',30)

        if self.cbo1.currentText() == '양산공장':
            B7 = char_format_left(number_register_Y, 10)  # 수정
        elif self.cbo1.currentText() == '대전공장':
            B7 = char_format_left(number_register_T,10) #수정
        else:
            B7 = char_format_left(number_register_W, 10)  # 수정


        B8 = char_format_left('1845110000543',13) #수정
        B9 = num_format_right('2019',4)
        B10 = num_format_right('1',1)

        B11 = num_format_right(number_of_people_v, 10)


        B12 = num_format_right(total_of_sal,13)
        B13 = num_format_right(total_of_bi_v,13)
        B14 = char_format_left(' ',44)


        record_B = B1+B2+B3+B4+B5+B6+B7+B8+B9+B10+B11+B12+B13+B14
        print(record_B)
        print(wcswidth(record_B))


        #파일에 쓰기  (A,B 레코드)

        if self.cbo1.currentText() == '양산공장':
            write = filename_Y
        elif self.cbo1.currentText() == '대전공장':
            write = filename_T
        else:
            write = filename_W

        f = open(write,'w')
        f.write(record_A)
        f.write('\n')
        f.write(record_B)
        f.write('\n')
        f. close()


        # C 레코드
        # 휴대폰 번호 (jpypf)

        if self.cbo1.currentText() == '양산공장':
            c1.execute(join_for_phone_Y)
        elif self.cbo1.currentText() == '대전공장':
            c1.execute(join_for_phone_T)
        else:
            c1.execute(join_for_phone_W)

        copy_phone_name=[]
        for i in c1:
            copy_phone_name.append(i[4])
    #        print(copy_phone_name)

        if self.cbo1.currentText() == '양산공장':
            c1.execute(select_all_q_Y)
        elif self.cbo1.currentText() == '대전공장':
            c1.execute(select_all_q_T)
        else:
            c1.execute(select_all_q_W)

        count_b =1
        k=0

        for i in c1:

            tmp = str(count_b)
            jumin = i[12]+i[13]
            name = i[11]

            if self.cbo2.currentText() == '전반기':

                gup = i[28] + i[37] + i[46] + i[55] + i[64] + i[73]
                sang = i[136] + i[141] + i[146] + i[151] + i[156] + i[161] + i[140] + i[145] + i[150] + i[155] + i[160] + i[165]
                total_gup = str(gup + sang)
                total_bi = str(i[36] + i[45] + i[54] + i[63] + i[72] + i[81])

                if int(i[4]) <= 20181231:
                    C12 = char_format_left('20190101', 8)
                else:
                    C12 = char_format_left(i[4], 8)

                if str(i[6]) =="        ":
                    C13 = char_format_left('20190630', 8)
                else:
                    C13 = char_format_left(i[6], 8)



            else:
                gup = i[82] + i[91] + i[100] + i[109] + i[118] + i[127]
                sang = i[166] + i[171] + i[176] + i[181] + i[186] + i[191] + i[170] + i[175] + i[180] + i[185] + i[190] + i[195]
                total_gup = str(gup + sang)
                total_bi = str(i[90] + i[99] + i[108] + i[117] + i[126] + i[135])

                if int(i[4]) <= 20190630:
                    C12 = char_format_left('20190701', 8)
                else:
                    C12 = char_format_left(i[4], 8)

                if str(i[6]) == "        ":
                    C13 = char_format_left('20191231', 8)
                else:
                    C13 = char_format_left(i[6], 8)


            print(i[6])


            C1 = char_format_left('C', 1)
            C2 = num_format_right('77', 2)



            if self.cbo1.currentText() == '양산공장':
                C3 = char_format_left(code_taxoffice_Y,3)
            elif self.cbo1.currentText() == '대전공장':
                C3 = char_format_left(code_taxoffice_T,3)
            else:
                C3 = char_format_left(code_taxoffice_W, 3)

            C4 = num_format_right(tmp,7)

            if self.cbo1.currentText() == '양산공장':
                C5 = char_format_left(number_register_Y,10)
            elif self.cbo1.currentText() == '대전공장':
                C5 = char_format_left(number_register_W,10)
            else:
                C5 = char_format_left(number_register_T, 10)


            C6 = char_format_left(jumin,13)
            C7 = char_format_left(name,30)
            C8 = char_format_left(copy_phone_name[k],20)   #copy_phone_name
            C9 = char_format_left('1',1)
            C10 = char_format_left('1',1)
            C11 = char_format_left('KR',2)





            C14 = num_format_right(total_gup,13)
            C15 = num_format_right(total_bi,13)
            C16 = char_format_left(' ', 58)

            record_C = C1+C2+C3+C4+C5+C6+C7+C8+C9+C10+C11+C12+C13+C14+C15+C16
            print(record_C)
            count_b += 1
            k+=1

            f = open(write,'a')
            f.write(record_C)
            f.write('\n')
            f. close()

        print(wcswidth(record_C))

        msg = ctypes.windll.user32.MessageBoxW(None, "완료", "파일추출", 0)
        if msg == 1:
            print("OK")
        connection.close()  # 연결 닫음



    def pushButtonClicked2(self):
        connection = pyodbc.connect(driver='{iSeries Access ODBC Driver}', system="192.168.1.3", uid='don2000',pwd='h2011')
        c1 = connection.cursor()

        wb = openpyxl.load_workbook('C:\\pay\\sample.xlsx')
        sheet1 = wb['sampleSheet']

        ft = Font('맑은고딕',size = 8)
        sheet1.font = ft

        sheet1.cell(row=6, column=2).value = "한일제관(주)"
        sheet1.cell(row=6, column=7).value = "정동택"

        if self.cbo1.currentText() == '양산공장':
            sheet1.cell(row=6, column=13).value = "6218105572"
            sheet1.cell(row=49, column=13).value = "6218105572"
        elif self.cbo1.currentText()=='음성공장':
            sheet1.cell(row=6, column=13).value = "1248531562"
            sheet1.cell(row=49, column=13).value = "1248531562"
        else:
            sheet1.cell(row=6, column=13).value = "3068512298"
            sheet1.cell(row=49, column=13).value = "3068512298"

        sheet1.cell(row=8, column=2).value = "1845110000543"

        if self.cbo1.currentText() == '양산공장':
            sheet1.cell(row=8, column=7).value = "경남 양산시 유산공단 4길 21"
        elif self.cbo1 == '음성공장':
            sheet1.cell(row=8, column=7).value = "충북 음성군 삼성면 하이텍산단로 84"
        else:
            sheet1.cell(row=8, column=7).value = "대전시 대덕구 대덕대로 1448번길 50"

        if self.cbo1.currentText() == '양산공장':
            sheet1.cell(row=10, column=2).value = "055-370-6600"
        elif self.cbo1 == '음성공장':
            sheet1.cell(row=10, column=2).value = "043-883-2805"
        else:
            sheet1.cell(row=10, column=2).value = "042-934-5500"


        sheet1.cell(row=10, column=7).value = "dtchung@hanilcan.co.kr"

        now = time.localtime()
        sheet1.cell(row=13, column=4).value = now.tm_year
        sheet1.cell(row=49, column=2).value = now.tm_year

        if self.cbo2.currentText() == "전반기":
           sheet1.cell(row=13, column=9).value = "[■]상반기(1월~6월)  [ ]하반기(7월~12월)"
           sheet1.cell(row=49, column=7).value = "[■]상반기  [ ]하반기"
        else :
            sheet1.cell(row=13, column=9).value = "[ ]상반기(1월~6월)  [■]하반기(7월~12월)"
            sheet1.cell(row=49, column=7).value = "[ ]상반기  [■]하반기"

        sheet1.cell(row=49, column=3).value = now.tm_year


        # 근로자 수 (전체인원)

        if  self.cbo1.currentText() == '양산공장':
            c1.execute(number_of_people_q_Y)
            tmp = c1.fetchone()
            number_of_people_v = str(tmp[0])
            sheet1.cell(row=16, column=3).value = str(number_of_people_v)
        elif self.cbo1.currentText() == '음성공장':
            c1.execute(number_of_people_q_W)
            tmp = c1.fetchone()
            number_of_people_v = str(tmp[0])
            sheet1.cell(row=16, column=3).value = str(number_of_people_v)
        else:
            c1.execute(number_of_people_q_T)
            tmp = c1.fetchone()
            number_of_people_v = str(tmp[0])
            sheet1.cell(row=16, column=3).value = str(number_of_people_v)



        # 과세소득 합계
        if self.cbo2.currentText() == "전반기":
        # 급여, 상여
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_gup_q_Y_6)
                sal1 = " ".join(re.findall("\d+", str(c1.fetchone())))
                c1.execute(total_of_sang_q_Y_6)
                sal2 = " ".join(re.findall("\d+", str(c1.fetchone())))


            elif self.cbo1.currentText() == '음성공장':
                c1.execute(total_of_gup_q_W_6)
                sal1 = " ".join(re.findall("\d+", str(c1.fetchone())))
                c1.execute(total_of_sang_q_W_6)
                sal2 = " ".join(re.findall("\d+", str(c1.fetchone())))

            else:
                c1.execute(total_of_gup_q_T_6)
                sal1 = " ".join(re.findall("\d+", str(c1.fetchone())))
                c1.execute(total_of_sang_q_W_6)
                sal2 = " ".join(re.findall("\d+", str(c1.fetchone())))

            sheet1.cell(row=16, column=8).value = str((int(sal1) + int(sal2)))



            # 비과세 소득 합계
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_bi_q_Y_6)

            elif self.cbo1.currentText() == '음성공장':
                c1.execute(total_of_bi_q_W_6)

            else:
                c1.execute(total_of_bi_q_T_6)

            tmp = " ".join(re.findall("\d+", str(c1.fetchone())))
            sheet1.cell(row=16, column=14).value = str(tmp)

        else:
            # 급여, 상여
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_gup_q_Y_12)
                sal1 = " ".join(re.findall("\d+", str(c1.fetchone())))
                c1.execute(total_of_sang_q_Y_12)
                sal2 = " ".join(re.findall("\d+", str(c1.fetchone())))


            elif self.cbo1.currentText() == '음성공장':
                c1.execute(total_of_gup_q_W_12)
                sal1 = " ".join(re.findall("\d+", str(c1.fetchone())))
                c1.execute(total_of_sang_q_W_12)
                sal2 = " ".join(re.findall("\d+", str(c1.fetchone())))

            else:
                c1.execute(total_of_gup_q_T_12)
                sal1 = " ".join(re.findall("\d+", str(c1.fetchone())))
                c1.execute(total_of_sang_q_T_12)
                sal2 = " ".join(re.findall("\d+", str(c1.fetchone())))

            sheet1.cell(row=16, column=8).value = str((int(sal1) + int(sal2)))

            # 비과세 소득 합계
            if self.cbo1.currentText() == '양산공장':
                c1.execute(total_of_bi_q_Y_12)

            elif self.cbo1.currentText() == '음성공장':
                c1.execute(total_of_bi_q_W_12)

            else:
                c1.execute(total_of_bi_q_T_12)

            tmp = " ".join(re.findall("\d+", str(c1.fetchone())))
            sheet1.cell(row=16, column=14).value = str(tmp)



        # 소득자 인적사항 및 근로소득 내용 (레코드 시작)


        if self.cbo1.currentText() == '양산공장':
            c1.execute(join_for_phone_Y)
        elif self.cbo1.currentText() == '대전공장':
            c1.execute(join_for_phone_T)
        else:
            c1.execute(join_for_phone_W)

        copy_phone_name = []
        for i in c1:
            copy_phone_name.append(i[4])


        if self.cbo1.currentText() == "양산공장":
            c=c1.execute(select_all_q_Y)
        elif self.cbo1.currentText() == "음성공장":
            c=c1.execute(select_all_q_W)
        else:
            c=c1.execute(select_all_q_T)



        rows = 23
        k=0



        for index,i in enumerate(c1):

            if self.cbo2.currentText() == '전반기':

                gup = i[28] + i[37] + i[46] + i[55] + i[64] + i[73]
                sang = i[136] + i[141] + i[146] + i[151] + i[156] + i[161] + i[140] + i[145] + i[150] + i[155] + i[160] + i[165]
                total_gup = str(gup + sang)
                total_bi = str(i[36] + i[45] + i[54] + i[63] + i[72] + i[81])

                if int(i[4]) <= 20181231:
                    year1 = '20190101-'
                else:
                    year1 = str(i[4]+ "-")

                if i[6] == "        ":
                    year2 = str(str(now.tm_year) + "0630")
                else:
                    year2 = str(i[6])

            else:
                gup = i[82] + i[91] + i[100] + i[109] + i[118] + i[127]
                sang = i[166] + i[171] + i[176] + i[181] + i[186] + i[191] + i[170] + i[175] + i[180] + i[185] + i[190] + i[195]
                total_gup = str(gup + sang)
                total_bi = str(i[90] + i[99] + i[108] + i[117] + i[126] + i[135])


                if int(i[4]) <= 20190630:
                    year1 = '20190701-'
                else:
                    year1 = str(i[4]+ "-")

                if i[6] == "        ":
                    year2 = str(str(now.tm_year) + "1231")
                else:
                    year2 = str(i[6])


            sheet1.cell(row=rows, column=3).value = i[12]+"-"
            sheet1.cell(row=rows, column=5).value = i[11]
            sheet1.cell(row=rows, column=8).value = "1"
            sheet1.cell(row=rows, column=10, value="korea,Republic of ").font = ft
            sheet1.cell(row=rows, column=12).value = year1
            sheet1.cell(row=rows, column=14).value = total_gup
            sheet1.cell(row=rows, column=16).value = total_bi

            rows = rows + 1

            sheet1.cell(row=rows, column=3).value = i[13]
            sheet1.cell(row=rows, column=5).value = str(copy_phone_name[k])
            sheet1.cell(row=rows, column=8).value = "1"
            sheet1.cell(row=rows, column=10).value = "KR"
            sheet1.cell(row=rows, column=12).value = year2
            sheet1.cell(row=rows, column=14).value = " "

            rows = rows + 1

            if rows == 37: rows = 55
            if rows == 83: rows = 91

            k=k+1


#            if int(number_of_people_v)  == sheet1.cell(row=o, column=3).value :
#               print(number_of_people_v)
#               p = int(number_of_people_v)
#               for o in range(p,300,1):
#                   print(o)
#                   sheet1.cell(row=o, column=3).value = None
#                   sheet1.cell(row=o, column=5).value = None
#                   sheet1.cell(row=o, column=8).value = None
#                   sheet1.cell(row=o, column=10).value = None
#                   sheet1.cell(row=o, column=12).value = None
#                   sheet1.cell(row=o, column=14).value = None
#                   sheet1.cell(row=o, column=16).value = None
#                   sheet1.cell(row=o, column=16).value = None



        # 뒤에 다 지우는 코드 넣을 것
        wb.save('C:\\pay\\sample.xlsx')
        #wb.save('C:\\Users\\cmjeong\\PycharmProjects\\exelTest\\sample.xlsx')
        print(222222222)

        msg = ctypes.windll.user32.MessageBoxW(None, "완료", "파일추출", 0)
        if msg == 1:
            print("OK")


        connection.close()

class MyApp(QMainWindow):

    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):

        menu = self.menuBar()                     # 메뉴바 생성

        menu_file = menu.addMenu('파일')          # 그룹 생성
        menu_sal = menu.addMenu('근로소득간이')   # 그룹 생성

        file_exit = QAction('종료',self)           # 메뉴 객체 생성
        file_exit.setShortcut('Ctrl+Q')
        file_exit.setStatusTip("누르면 종료")
        file_exit.triggered.connect(QCoreApplication.instance().quit)

        cal_sal = QAction('파일추출', self)
        cal_sal.triggered.connect(self.open_new_window)

    #    file_new = QMenu('New',self)                # 서브그룹
    #    menu_file.addMenu(file_new)                 #서브 메뉴 추가

        menu_file.addAction(file_exit)              # 메뉴 등록
        menu_sal.addAction(cal_sal)

        self.resize(450,400)
        self.show()



    def open_new_window(self):
        dlg = LogInDialog1()
        dlg.exec_()


if __name__ == '__main__':

    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())